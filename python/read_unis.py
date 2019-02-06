from openpyxl import load_workbook
import json
import re
from bson import json_util
import unicodedata


def standardize(name):
    name = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
    name = re.sub(' ', '_', name).lower().strip().replace("\n", "")

    return name


def clean_uni_str(string):
    """Removed brackets and its content from string."""
    reg_brackets = re.compile('([\w*\s*\'’-]*)[\(,]?([\w*\s*]*)', re.UNICODE)  # Find words up until left bracket
    reg_star = re.compile("\*")

    clean_str = re.sub(reg_star, '', string)  # Remove star
    matches = re.match(reg_brackets, clean_str)  # Strip content in brackets or after comma

    return matches.group(1).strip()


def main():
    """Script to read universities from the Excel sheet provided by LTH of universities that are open for application.
    """
    # Read workbook from excel file in same directory
    wb = load_workbook('./01Engineering_ht18_web.xlsx')

    # Get first and only sheet
    sheet = wb[wb.sheetnames[0]]

    # Find row of headers
    header_row = 0

    for row in range(1, sheet.max_row):
        cell = sheet.cell(row=row, column=1).value
        if cell is not None and cell.lower() == 'country':
            header_row = row
            break

    # Traverse whole sheet and add to JSON
    JSON = {}
    c_col = 1
    u_col = 2

    # define fieldnames
    pretty = "pretty"
    unis = "unis"

    for row in range(header_row + 1, sheet.max_row + 1):
        country = sheet.cell(row=row, column=c_col).value
        uni = sheet.cell(row=row, column=u_col).value

        # Stop traversing if country is empty and ignore middle header rows
        if country is None:
            break
        elif country.lower() != 'country':

            country_pretty = country
            country = standardize(country_pretty)
            uni_pretty = clean_uni_str(uni)
            uni = standardize(uni_pretty)

            if country is not None and country in JSON:  # Country has been added to JSON
                if uni not in JSON[country][unis]:  # Only add if not already present
                    JSON[country][unis][uni] = uni_pretty
            else:
                JSON[country] = {unis: {uni: uni_pretty}, pretty: country_pretty, }

    with open('data.json', 'w', encoding='utf-8') as file:
        json.dump(JSON, file, ensure_ascii=False, default=json_util.default)

        print("Saved data in [data.json]")


if __name__ == '__main__':
    main()
